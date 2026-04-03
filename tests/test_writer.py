"""Tests for writer module."""

from pathlib import Path

import openpyxl
import pytest

from excel_mcp.writer.formatter import add_conditional_format, format_range
from excel_mcp.writer.templates import open_template
from excel_mcp.writer.verify import verify_workbook
from excel_mcp.writer.writer import create_table, create_workbook, write_data, write_formula

FIXTURES = Path(__file__).parent / "fixtures"


class TestCreateWorkbook:
    def test_single_sheet(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        result = create_workbook(path)
        assert result["path"] == path
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Sheet1"

        # Verify file is valid
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Sheet1"]
        wb.close()

    def test_multiple_sheets(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        result = create_workbook(path, sheets=["Data", "Summary", "Notes"])
        assert len(result["sheets"]) == 3
        names = [s["name"] for s in result["sheets"]]
        assert names == ["Data", "Summary", "Notes"]

    def test_creates_parent_dirs(self, tmp_path):
        path = str(tmp_path / "deep" / "nested" / "test.xlsx")
        create_workbook(path)
        assert Path(path).exists()


class TestWriteData:
    def test_basic_write(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)

        data = [
            ["Name", "Age", "City"],
            ["Alice", 30, "London"],
            ["Bob", 25, "Paris"],
        ]
        result = write_data(path, "Sheet1", "A1", data)
        assert result["rows_written"] == 3
        assert result["cols_written"] == 3

        # Verify
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws["A1"].value == "Name"
        assert ws["B2"].value == 30
        assert ws["C3"].value == "Paris"
        wb.close()

    def test_write_at_offset(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)

        data = [["X", "Y"], [1, 2]]
        write_data(path, "Sheet1", "C5", data)

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws["C5"].value == "X"
        assert ws["D6"].value == 2
        wb.close()

    def test_large_write(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)

        data = [[f"row{i}", i, i * 1.5] for i in range(1000)]
        result = write_data(path, "Sheet1", "A1", data)
        assert result["rows_written"] == 1000

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(row=1000, column=1).value == "row999"
        wb.close()


class TestWriteFormula:
    def test_basic_formula(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [[10], [20], [30]])

        result = write_formula(path, "Sheet1", "A4", "=SUM(A1:A3)")
        assert result["formula"] == "=SUM(A1:A3)"

        wb = openpyxl.load_workbook(path, data_only=False)
        assert wb.active["A4"].value == "=SUM(A1:A3)"
        wb.close()

    def test_formula_with_format(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)

        write_formula(path, "Sheet1", "B1", "=100/3", number_format="#,##0.00")

        wb = openpyxl.load_workbook(path)
        assert wb.active["B1"].number_format == "#,##0.00"
        wb.close()

    def test_auto_prefix_equals(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)

        result = write_formula(path, "Sheet1", "A1", "SUM(A2:A10)")
        assert result["formula"] == "=SUM(A2:A10)"


class TestFormatRange:
    def test_bold_font(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [["Header1", "Header2"]])

        result = format_range(path, "Sheet1", "A1:B1", {"font": {"bold": True, "size": 14}})
        assert result["cells_formatted"] == 2

        wb = openpyxl.load_workbook(path)
        assert wb.active["A1"].font.bold is True
        assert wb.active["A1"].font.size == 14
        wb.close()

    def test_fill_and_border(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [["Data"]])

        format_range(path, "Sheet1", "A1", {
            "fill": {"color": "FFFF00"},
            "border": "thin",
        })

        wb = openpyxl.load_workbook(path)
        cell = wb.active["A1"]
        assert cell.fill.fgColor.rgb == "00FFFF00"
        assert cell.border.left.style == "thin"
        wb.close()

    def test_number_format(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [[1234.5678]])

        format_range(path, "Sheet1", "A1", {"number_format": "#,##0.00"})

        wb = openpyxl.load_workbook(path)
        assert wb.active["A1"].number_format == "#,##0.00"
        wb.close()

    def test_alignment(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [["Centered"]])

        format_range(path, "Sheet1", "A1", {
            "alignment": {"horizontal": "center", "wrap_text": True},
        })

        wb = openpyxl.load_workbook(path)
        assert wb.active["A1"].alignment.horizontal == "center"
        assert wb.active["A1"].alignment.wrap_text is True
        wb.close()


class TestCreateTable:
    def test_basic_table(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [
            ["Name", "Value"],
            ["A", 10],
            ["B", 20],
        ])

        result = create_table(path, "Sheet1", "A1:B3", "TestTable")
        assert result["table_name"] == "TestTable"

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert "TestTable" in ws.tables
        wb.close()

    def test_table_with_totals(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [
            ["Name", "Value"],
            ["A", 10],
            ["B", 20],
        ])

        result = create_table(
            path, "Sheet1", "A1:B3", "TotalsTable", totals_row=True
        )
        assert result["totals_row"] is True


class TestConditionalFormat:
    def test_cell_is_rule(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [[10], [-5], [20], [-15]])

        result = add_conditional_format(path, "Sheet1", "A1:A4", {
            "type": "cell_is",
            "operator": "lessThan",
            "value": 0,
            "font": {"color": "FF0000"},
        })
        assert result["rule_type"] == "cell_is"

    def test_color_scale(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [[i] for i in range(10)])

        result = add_conditional_format(path, "Sheet1", "A1:A10", {
            "type": "color_scale",
            "start_color": "FF0000",
            "end_color": "00FF00",
        })
        assert result["rule_type"] == "color_scale"

    def test_data_bar(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [[i * 10] for i in range(5)])

        result = add_conditional_format(path, "Sheet1", "A1:A5", {
            "type": "data_bar",
            "color": "638EC6",
        })
        assert result["rule_type"] == "data_bar"


class TestOpenTemplate:
    def test_copy_template(self, tmp_path):
        # Use simple_table.xlsx as a "template"
        template = str(FIXTURES / "simple_table.xlsx")
        output = str(tmp_path / "output.xlsx")

        result = open_template(template, output)
        assert result["template"] == template
        assert result["output"] == output
        assert Path(output).exists()
        assert any(s["name"] == "Employees" for s in result["sheets"])

    def test_preserves_data(self, tmp_path):
        template = str(FIXTURES / "simple_table.xlsx")
        output = str(tmp_path / "output.xlsx")
        open_template(template, output)

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws["A1"].value == "Name"
        assert ws["A2"].value == "Alice"
        wb.close()

    def test_template_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            open_template("/nonexistent/template.xlsx", str(tmp_path / "out.xlsx"))


class TestVerifyWorkbook:
    def test_all_checks_pass(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path, sheets=["Data"])
        write_data(path, "Data", "A1", [["X"], [1], [2], [3]])

        result = verify_workbook(path, [
            {"type": "sheet_exists", "sheet": "Data"},
            {"type": "row_count", "sheet": "Data", "expected": 4},
            {"type": "cell_value", "sheet": "Data", "cell": "A2", "expected": 1},
            {"type": "file_size_min", "min_bytes": 100},
        ])
        assert result["valid"] is True
        assert all(r["passed"] for r in result["results"])

    def test_row_count_mismatch(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [["A"], ["B"]])

        result = verify_workbook(path, [
            {"type": "row_count", "sheet": "Sheet1", "expected": 99},
        ])
        assert result["valid"] is False

    def test_cell_value_tolerance(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        create_workbook(path)
        write_data(path, "Sheet1", "A1", [[100.005]])

        result = verify_workbook(path, [
            {"type": "cell_value", "sheet": "Sheet1", "cell": "A1", "expected": 100.01},
        ])
        assert result["valid"] is True

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            verify_workbook("/nonexistent/file.xlsx", [])
