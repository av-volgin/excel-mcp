"""Test fixtures — generate synthetic xlsx files for testing."""

from pathlib import Path

import openpyxl
import pytest


FIXTURES_DIR = Path(__file__).parent / "fixtures"


@pytest.fixture(scope="session", autouse=True)
def create_fixtures():
    """Generate test xlsx files once per test session."""
    FIXTURES_DIR.mkdir(exist_ok=True)
    _create_simple_table()
    _create_formulas_file()
    _create_multi_sheet()


def _create_simple_table():
    """Simple table: headers + data rows + total row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = ["Name", "Department", "Salary", "Start Date"]
    ws.append(headers)

    data = [
        ["Alice", "Engineering", 95000, "2023-01-15"],
        ["Bob", "Marketing", 72000, "2023-03-20"],
        ["Charlie", "Engineering", 88000, "2022-11-01"],
        ["Diana", "Sales", 67000, "2024-02-10"],
        ["Eve", "Engineering", 102000, "2021-06-15"],
    ]
    for row in data:
        ws.append(row)

    # Total row (bold)
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=1).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=total_row, column=3, value=424000)
    ws.cell(row=total_row, column=3).font = openpyxl.styles.Font(bold=True)

    wb.save(FIXTURES_DIR / "simple_table.xlsx")
    wb.close()


def _create_formulas_file():
    """File with formulas for testing formula extraction."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"

    ws["A1"] = "Month"
    ws["B1"] = "Income"
    ws["C1"] = "Expenses"
    ws["D1"] = "Profit"

    months = ["Jan", "Feb", "Mar"]
    incomes = [50000, 62000, 58000]
    expenses = [35000, 41000, 37000]

    for i, (month, inc, exp) in enumerate(zip(months, incomes, expenses), start=2):
        ws[f"A{i}"] = month
        ws[f"B{i}"] = inc
        ws[f"C{i}"] = exp
        ws[f"D{i}"] = f"=B{i}-C{i}"

    # Totals with SUM formulas
    ws["A5"] = "Total"
    ws["B5"] = "=SUM(B2:B4)"
    ws["C5"] = "=SUM(C2:C4)"
    ws["D5"] = "=SUM(D2:D4)"

    # Summary cell
    ws["A7"] = "Margin %"
    ws["B7"] = "=D5/B5*100"

    wb.save(FIXTURES_DIR / "formulas.xlsx")
    wb.close()


def _create_multi_sheet():
    """Workbook with multiple sheets of different types."""
    wb = openpyxl.Workbook()

    # Sheet 1: data table
    ws1 = wb.active
    ws1.title = "Transactions"
    ws1.append(["Date", "Description", "Amount", "Category"])
    ws1.append(["2026-01-15", "Office rent", 50000, "Operations"])
    ws1.append(["2026-01-16", "Software license", 12000, "IT"])
    ws1.append(["2026-01-20", "Salary payment", 180000, "Personnel"])
    ws1.append(["2026-01-25", "Client payment", 250000, "Revenue"])

    # Sheet 2: summary with formulas
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Metric"
    ws2["B1"] = "Value"
    ws2["A2"] = "Total Expenses"
    ws2["B2"] = 242000
    ws2["A3"] = "Total Revenue"
    ws2["B3"] = 250000
    ws2["A4"] = "Net"
    ws2["B4"] = "=B3-B2"

    # Sheet 3: empty sheet
    wb.create_sheet("Notes")

    wb.save(FIXTURES_DIR / "multi_sheet.xlsx")
    wb.close()
