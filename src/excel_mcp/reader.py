"""Excel reading functions using openpyxl (bulk data extraction)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from .utils import classify_value, coords_to_cell_ref, is_total_row, parse_range


def _open_workbook(file_path: str, data_only: bool = True) -> openpyxl.Workbook:
    """Open workbook with validation."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    if path.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise ValueError(f"Unsupported file format: {path.suffix}")
    return openpyxl.load_workbook(str(path), data_only=data_only, read_only=True)


def list_sheets(file_path: str) -> list[dict[str, Any]]:
    """List all sheets in a workbook with metadata.

    Returns list of {name, rows, cols, has_tables, table_names}.
    """
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    try:
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            tables = list(ws.tables.keys()) if hasattr(ws, "tables") else []
            sheets.append({
                "name": name,
                "rows": ws.max_row or 0,
                "cols": ws.max_column or 0,
                "has_tables": len(tables) > 0,
                "table_names": tables,
            })
        return sheets
    finally:
        wb.close()


def read_sheet(
    file_path: str,
    sheet_name: str | None = None,
    header_row: int | None = None,
    start_row: int | None = None,
    max_rows: int = 500,
    columns: list[str] | None = None,
    skip_totals: bool = True,
) -> dict[str, Any]:
    """Read a sheet as a table with headers and rows.

    Args:
        file_path: Path to xlsx file.
        sheet_name: Sheet name (default: first sheet).
        header_row: Row number with headers (default: auto-detect).
        start_row: First data row (default: header_row + 1).
        max_rows: Maximum rows to return (pagination).
        columns: Filter to specific column names.
        skip_totals: Skip rows that look like totals (bold, "Итого"/"Total").

    Returns:
        {headers, rows, total_rows, has_more, sheet_name}
    """
    wb = _open_workbook(file_path, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        actual_sheet = ws.title

        # Auto-detect header row: first row with >=2 non-empty cells
        if header_row is None:
            header_row = _detect_header_row(ws)

        # Read headers
        headers = []
        for col_idx in range(1, (ws.max_column or 0) + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val is not None:
                headers.append(str(val).strip())
            else:
                headers.append(get_column_letter(col_idx))

        if start_row is None:
            start_row = header_row + 1

        # Filter column indices if specified
        col_indices = None
        if columns:
            col_indices = []
            for cn in columns:
                if cn in headers:
                    col_indices.append(headers.index(cn))

        # Read data rows
        rows = []
        total_rows = 0
        max_row = ws.max_row or 0

        for row_idx in range(start_row, max_row + 1):
            row_values = [ws.cell(row=row_idx, column=c + 1).value for c in range(len(headers))]

            # Skip completely empty rows
            if all(v is None for v in row_values):
                continue

            total_rows += 1

            # Skip total rows if requested
            if skip_totals and is_total_row(row_values):
                continue

            if len(rows) < max_rows:
                row_dict = {}
                for i, h in enumerate(headers):
                    if col_indices is not None and i not in col_indices:
                        continue
                    val, _ = classify_value(row_values[i])
                    row_dict[h] = val
                rows.append(row_dict)

        return {
            "headers": headers if col_indices is None else [headers[i] for i in col_indices],
            "rows": rows,
            "total_rows": total_rows,
            "has_more": total_rows > len(rows),
            "sheet_name": actual_sheet,
        }
    finally:
        wb.close()


def read_cell_range(
    file_path: str,
    sheet_name: str,
    range_str: str,
) -> dict[str, Any]:
    """Read raw cell values from a range (e.g. 'A1:D10' or 'B5').

    Returns {cells: [{ref, row, col, value, type}], sheet_name}.
    """
    wb = _open_workbook(file_path, data_only=True)
    try:
        ws = wb[sheet_name]
        (min_row, min_col), (max_row, max_col) = parse_range(range_str)

        cells = []
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                raw = ws.cell(row=r, column=c).value
                val, vtype = classify_value(raw)
                cells.append({
                    "ref": coords_to_cell_ref(r, c),
                    "row": r,
                    "col": c,
                    "value": val,
                    "type": vtype,
                })
        return {
            "cells": cells,
            "sheet_name": sheet_name,
        }
    finally:
        wb.close()


def _detect_header_row(ws: Any, max_scan: int = 20) -> int:
    """Find the first row with at least 2 non-empty cells."""
    max_col = ws.max_column or 1
    for row_idx in range(1, min((ws.max_row or 1) + 1, max_scan + 1)):
        non_empty = sum(
            1 for c in range(1, max_col + 1) if ws.cell(row=row_idx, column=c).value is not None
        )
        if non_empty >= 2:
            return row_idx
    return 1
