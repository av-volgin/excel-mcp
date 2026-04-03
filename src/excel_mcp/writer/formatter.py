"""Styling and conditional formatting for Excel files."""

from __future__ import annotations

from typing import Any

from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..common.utils import parse_range
from ..common.workbook import open_for_write

# Border style mapping
BORDER_STYLES = {
    "thin": Side(style="thin"),
    "medium": Side(style="medium"),
    "thick": Side(style="thick"),
    "dashed": Side(style="dashed"),
    "dotted": Side(style="dotted"),
    "double": Side(style="double"),
    "none": Side(style=None),
}


def format_range(
    file_path: str,
    sheet_name: str,
    range_str: str,
    styles: dict[str, Any],
) -> dict[str, Any]:
    """Apply formatting to a range of cells.

    Args:
        file_path: Absolute path to the .xlsx file.
        sheet_name: Target sheet name.
        range_str: Cell range, e.g. "A1:D10".
        styles: Style dictionary with optional keys:
            font: {bold, italic, underline, size, color, name}
            fill: {color, type}
            border: "thin" or {left, right, top, bottom}
            number_format: "#,##0.00"
            alignment: {horizontal, vertical, wrap_text}

    Returns:
        {cells_formatted, range}
    """
    wb = open_for_write(file_path)
    try:
        ws = wb[sheet_name]
        (min_row, min_col), (max_row, max_col) = parse_range(range_str)

        font_style = _build_font(styles.get("font")) if "font" in styles else None
        fill_style = _build_fill(styles.get("fill")) if "fill" in styles else None
        border_style = _build_border(styles.get("border")) if "border" in styles else None
        alignment_style = (
            _build_alignment(styles.get("alignment")) if "alignment" in styles else None
        )
        number_format = styles.get("number_format")

        count = 0
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if font_style:
                    cell.font = font_style
                if fill_style:
                    cell.fill = fill_style
                if border_style:
                    cell.border = border_style
                if alignment_style:
                    cell.alignment = alignment_style
                if number_format:
                    cell.number_format = number_format
                count += 1

        wb.save(file_path)
        return {"cells_formatted": count, "range": range_str}
    finally:
        wb.close()


def add_conditional_format(
    file_path: str,
    sheet_name: str,
    range_str: str,
    rule: dict[str, Any],
) -> dict[str, Any]:
    """Add conditional formatting to a range.

    Args:
        file_path: Absolute path to the .xlsx file.
        sheet_name: Target sheet name.
        range_str: Cell range, e.g. "C2:C100".
        rule: Rule definition with 'type' and type-specific params:
            type: "cell_is" — {operator, value, font?, fill?}
            type: "color_scale" — {start_color, end_color, mid_color?}
            type: "data_bar" — {color}

    Returns:
        {rule_type, range}
    """
    wb = open_for_write(file_path)
    try:
        ws = wb[sheet_name]
        rule_type = rule["type"]

        if rule_type == "cell_is":
            cf_rule = CellIsRule(
                operator=rule.get("operator", "lessThan"),
                formula=[str(rule.get("value", 0))],
                font=_build_font(rule.get("font")) if "font" in rule else None,
                fill=_build_fill(rule.get("fill")) if "fill" in rule else None,
            )
            ws.conditional_formatting.add(range_str, cf_rule)

        elif rule_type == "color_scale":
            cf_rule = ColorScaleRule(
                start_type="min",
                start_color=rule.get("start_color", "FF0000"),
                mid_type="percentile" if "mid_color" in rule else None,
                mid_value=50 if "mid_color" in rule else None,
                mid_color=rule.get("mid_color"),
                end_type="max",
                end_color=rule.get("end_color", "00FF00"),
            )
            ws.conditional_formatting.add(range_str, cf_rule)

        elif rule_type == "data_bar":
            cf_rule = DataBarRule(
                start_type="min",
                end_type="max",
                color=rule.get("color", "638EC6"),
            )
            ws.conditional_formatting.add(range_str, cf_rule)

        else:
            raise ValueError(f"Unknown conditional format rule type: {rule_type}")

        wb.save(file_path)
        return {"rule_type": rule_type, "range": range_str}
    finally:
        wb.close()


def _build_font(font_dict: dict[str, Any] | None) -> Font | None:
    """Build openpyxl Font from a dict."""
    if not font_dict:
        return None
    return Font(
        bold=font_dict.get("bold", False),
        italic=font_dict.get("italic", False),
        underline=font_dict.get("underline"),
        size=font_dict.get("size"),
        color=font_dict.get("color"),
        name=font_dict.get("name"),
    )


def _build_fill(fill_dict: dict[str, Any] | None) -> PatternFill | None:
    """Build openpyxl PatternFill from a dict."""
    if not fill_dict:
        return None
    return PatternFill(
        patternType=fill_dict.get("type", "solid"),
        fgColor=fill_dict.get("color", "FFFFFF"),
    )


def _build_border(border_spec: str | dict[str, str] | None) -> Border | None:
    """Build openpyxl Border from a string or dict."""
    if not border_spec:
        return None

    if isinstance(border_spec, str):
        side = BORDER_STYLES.get(border_spec, Side(style=border_spec))
        return Border(left=side, right=side, top=side, bottom=side)

    return Border(
        left=BORDER_STYLES.get(border_spec.get("left", "none"), Side()),
        right=BORDER_STYLES.get(border_spec.get("right", "none"), Side()),
        top=BORDER_STYLES.get(border_spec.get("top", "none"), Side()),
        bottom=BORDER_STYLES.get(border_spec.get("bottom", "none"), Side()),
    )


def _build_alignment(align_dict: dict[str, Any] | None) -> Alignment | None:
    """Build openpyxl Alignment from a dict."""
    if not align_dict:
        return None
    return Alignment(
        horizontal=align_dict.get("horizontal"),
        vertical=align_dict.get("vertical"),
        wrap_text=align_dict.get("wrap_text", False),
    )
