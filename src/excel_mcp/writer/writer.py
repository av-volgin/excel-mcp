"""Core write logic — create workbooks, write data and formulas."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from ..common.utils import cell_ref_to_coords, parse_range
from ..common.workbook import open_for_write


def create_workbook(file_path: str, sheets: list[str] | None = None) -> dict[str, Any]:
    """Create a new .xlsx workbook with specified sheets.

    Args:
        file_path: Absolute path for the new file.
        sheets: List of sheet names (default: ["Sheet1"]).

    Returns:
        {path, sheets: [{name, rows, cols}]}
    """
    if sheets is None:
        sheets = ["Sheet1"]

    wb = openpyxl.Workbook()

    # Rename default sheet or remove it
    default_ws = wb.active
    if sheets:
        default_ws.title = sheets[0]
        for name in sheets[1:]:
            wb.create_sheet(name)
    else:
        wb.remove(default_ws)

    # Ensure parent directory exists
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    wb.close()

    return {
        "path": file_path,
        "sheets": [{"name": s, "rows": 0, "cols": 0} for s in sheets],
    }


def write_data(
    file_path: str,
    sheet_name: str,
    range_str: str,
    data: list[list[Any]],
) -> dict[str, Any]:
    """Write a 2D array of values into a cell range.

    Args:
        file_path: Absolute path to the .xlsx file.
        sheet_name: Target sheet name.
        range_str: Starting cell or range, e.g. "A1" or "B2:G100".
        data: 2D array of values (list of rows).

    Returns:
        {rows_written, cols_written, range}
    """
    wb = open_for_write(file_path)
    try:
        ws = wb[sheet_name]
        (start_row, start_col), _ = parse_range(range_str)

        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)

        rows_written = len(data)
        cols_written = max(len(row) for row in data) if data else 0

        wb.save(file_path)
        return {
            "rows_written": rows_written,
            "cols_written": cols_written,
            "range": range_str,
        }
    finally:
        wb.close()


def write_formula(
    file_path: str,
    sheet_name: str,
    cell: str,
    formula: str,
    number_format: str | None = None,
) -> dict[str, Any]:
    """Write a formula to a specific cell.

    Args:
        file_path: Absolute path to the .xlsx file.
        sheet_name: Target sheet name.
        cell: Cell reference, e.g. "D5".
        formula: Excel formula string, e.g. "=SUM(A1:A10)".
        number_format: Optional number format, e.g. "#,##0.00".

    Returns:
        {cell, formula}
    """
    wb = open_for_write(file_path)
    try:
        ws = wb[sheet_name]
        row, col = cell_ref_to_coords(cell)
        ws_cell = ws.cell(row=row, column=col)

        # Ensure formula starts with =
        if not formula.startswith("="):
            formula = "=" + formula

        ws_cell.value = formula

        if number_format:
            ws_cell.number_format = number_format

        wb.save(file_path)
        return {"cell": cell, "formula": formula}
    finally:
        wb.close()


def create_table(
    file_path: str,
    sheet_name: str,
    range_str: str,
    table_name: str,
    style: str = "TableStyleMedium2",
    totals_row: bool = False,
) -> dict[str, Any]:
    """Create an Excel Table (ListObject) from a data range.

    Args:
        file_path: Absolute path to the .xlsx file.
        sheet_name: Sheet containing the data.
        range_str: Data range including headers, e.g. "A1:D20".
        table_name: Display name for the table.
        style: Table style name (default: TableStyleMedium2).
        totals_row: Whether to add a totals row.

    Returns:
        {table_name, range, style, totals_row}
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = open_for_write(file_path)
    try:
        ws = wb[sheet_name]

        table_style = TableStyleInfo(
            name=style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table = Table(
            displayName=table_name,
            ref=range_str,
        )
        table.tableStyleInfo = table_style

        if totals_row:
            table.totalsRowShown = True
            table.totalsRowCount = 1

        ws.add_table(table)
        wb.save(file_path)

        return {
            "table_name": table_name,
            "range": range_str,
            "style": style,
            "totals_row": totals_row,
        }
    finally:
        wb.close()
