"""Formula-aware Excel analysis using Formualizer (with openpyxl fallback)."""

from __future__ import annotations

from typing import Any

import openpyxl

from ..common.utils import cell_ref_to_coords, classify_value, coords_to_cell_ref, parse_range
from ..common.workbook import open_for_read

# Formualizer is optional
try:
    import formualizer

    HAS_FORMUALIZER = True
except ImportError:
    HAS_FORMUALIZER = False


def get_formulas(
    file_path: str,
    sheet_name: str | None = None,
    cell_range: str | None = None,
) -> dict[str, Any]:
    """Extract formulas from a sheet or cell range.

    With Formualizer: returns formula text + evaluated values.
    Without Formualizer: returns formula text only (openpyxl).

    Returns {formulas: [{cell, formula_text, evaluated_value?}], engine, sheet_name}.
    """
    from pathlib import Path

    if not Path(file_path).exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if HAS_FORMUALIZER:
        return _get_formulas_formualizer(file_path, sheet_name, cell_range)
    return _get_formulas_openpyxl(file_path, sheet_name, cell_range)


def get_cell_value(
    file_path: str,
    sheet_name: str,
    row: int,
    col: int,
) -> dict[str, Any]:
    """Read a single cell value by coordinates.

    With Formualizer: fast targeted read via get_value.
    Without Formualizer: openpyxl data_only=True.

    Returns {ref, value, type, formula?, evaluated_value?, engine}.
    """
    from pathlib import Path

    if not Path(file_path).exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    ref = coords_to_cell_ref(row, col)

    if HAS_FORMUALIZER:
        return _get_cell_formualizer(file_path, sheet_name, row, col, ref)
    return _get_cell_openpyxl(file_path, sheet_name, row, col, ref)


def validate_totals(
    file_path: str,
    sheet_name: str,
    checks: list[dict[str, Any]],
) -> dict[str, Any]:
    """Validate cell values against expected values.

    Args:
        checks: List of {cell: "B10", expected_value: 12345.67}

    Returns {valid: bool, results: [{cell, expected, actual, match}]}.
    """
    from pathlib import Path

    if not Path(file_path).exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    engine = "formualizer" if HAS_FORMUALIZER else "openpyxl"
    results = []
    all_match = True

    for check in checks:
        cell_ref = check["cell"]
        expected = check["expected_value"]
        row, col = cell_ref_to_coords(cell_ref)

        if HAS_FORMUALIZER:
            wb = formualizer.load_workbook(file_path)
            actual = wb.get_value(sheet_name, row, col)  # 1-based in v0.5.4+
        else:
            wb_opx = open_for_read(file_path, data_only=True)
            try:
                ws = wb_opx[sheet_name]
                actual = ws.cell(row=row, column=col).value
            finally:
                wb_opx.close()

        # Compare with tolerance for floats
        if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
            match = abs(expected - actual) < 0.01
        else:
            match = expected == actual

        if not match:
            all_match = False

        actual_ser, _ = classify_value(actual)
        results.append({
            "cell": cell_ref,
            "expected": expected,
            "actual": actual_ser,
            "match": match,
        })

    return {
        "valid": all_match,
        "results": results,
        "engine": engine,
    }


# --- Formualizer implementations ---


def _get_formulas_formualizer(
    file_path: str,
    sheet_name: str | None,
    cell_range: str | None,
) -> dict[str, Any]:
    """Extract formulas using Formualizer (text + evaluated values)."""
    wb = formualizer.load_workbook(file_path)

    # Use openpyxl for sheet names (formualizer may include phantom sheets)
    wb_opx = openpyxl.load_workbook(file_path, read_only=True)
    try:
        if sheet_name is None:
            sheet_name = wb_opx.sheetnames[0]

        # Determine scan range
        if cell_range:
            (min_row, min_col), (max_row, max_col) = parse_range(cell_range)
        else:
            ws = wb_opx[sheet_name]
            min_row, min_col = 1, 1
            max_row = ws.max_row or 100
            max_col = ws.max_column or 26
    finally:
        wb_opx.close()

    formulas = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            try:
                formula_text = wb.get_formula(sheet_name, r, c)  # 1-based in v0.5.4+
            except Exception:
                formula_text = None

            if formula_text:
                entry: dict[str, Any] = {
                    "cell": coords_to_cell_ref(r, c),
                    "formula_text": formula_text,
                }
                # Evaluate formula value
                try:
                    evaluated = wb.evaluate_cell(sheet_name, r, c)
                    val, _ = classify_value(evaluated)
                    entry["evaluated_value"] = val
                except Exception:
                    entry["evaluated_value"] = None

                formulas.append(entry)

    return {
        "formulas": formulas,
        "engine": "formualizer",
        "sheet_name": sheet_name,
    }


def _get_formulas_openpyxl(
    file_path: str,
    sheet_name: str | None,
    cell_range: str | None,
) -> dict[str, Any]:
    """Extract formulas using openpyxl (text only, no AST)."""
    wb = open_for_read(file_path, data_only=False)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        actual_sheet = ws.title

        if cell_range:
            (min_row, min_col), (max_row, max_col) = parse_range(cell_range)
        else:
            min_row, min_col = 1, 1
            max_row = ws.max_row or 100
            max_col = ws.max_column or 26

        formulas = []
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and val.startswith("="):
                    formulas.append({
                        "cell": coords_to_cell_ref(r, c),
                        "formula_text": val,
                    })

        return {
            "formulas": formulas,
            "engine": "openpyxl",
            "sheet_name": actual_sheet,
        }
    finally:
        wb.close()


def _get_cell_formualizer(
    file_path: str,
    sheet_name: str,
    row: int,
    col: int,
    ref: str,
) -> dict[str, Any]:
    """Read single cell via Formualizer."""
    wb = formualizer.load_workbook(file_path)
    value = wb.get_value(sheet_name, row, col)  # 1-based in v0.5.4+
    val, vtype = classify_value(value)

    result: dict[str, Any] = {
        "ref": ref,
        "value": val,
        "type": vtype,
        "engine": "formualizer",
    }

    try:
        formula = wb.get_formula(sheet_name, row, col)
        if formula:
            result["formula"] = formula
            try:
                evaluated = wb.evaluate_cell(sheet_name, row, col)
                ev_val, _ = classify_value(evaluated)
                result["evaluated_value"] = ev_val
            except Exception:
                pass
    except Exception:
        pass

    return result


def _get_cell_openpyxl(
    file_path: str,
    sheet_name: str,
    row: int,
    col: int,
    ref: str,
) -> dict[str, Any]:
    """Read single cell via openpyxl."""
    # Read value
    wb_val = open_for_read(file_path, data_only=True)
    try:
        value = wb_val[sheet_name].cell(row=row, column=col).value
    finally:
        wb_val.close()

    val, vtype = classify_value(value)
    result: dict[str, Any] = {
        "ref": ref,
        "value": val,
        "type": vtype,
        "engine": "openpyxl",
    }

    # Read formula
    wb_formula = open_for_read(file_path, data_only=False)
    try:
        formula_val = wb_formula[sheet_name].cell(row=row, column=col).value
        if isinstance(formula_val, str) and formula_val.startswith("="):
            result["formula"] = formula_val
    finally:
        wb_formula.close()

    return result
