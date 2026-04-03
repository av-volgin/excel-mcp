"""Shared workbook loading and saving utilities."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from .errors import FileValidationError

SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm")


def open_for_read(file_path: str, data_only: bool = True) -> openpyxl.Workbook:
    """Open workbook in read-only mode with validation.

    Args:
        file_path: Absolute path to the .xlsx file.
        data_only: If True, return cached values instead of formulas.
    """
    _validate_file(file_path)
    return openpyxl.load_workbook(str(file_path), data_only=data_only, read_only=True)


def open_for_write(file_path: str) -> openpyxl.Workbook:
    """Open existing workbook for modification (read-write mode).

    Args:
        file_path: Absolute path to the .xlsx file.
    """
    _validate_file(file_path)
    return openpyxl.load_workbook(str(file_path), data_only=False)


def open_for_metadata(file_path: str) -> openpyxl.Workbook:
    """Open workbook in full mode for metadata access (tables, etc).

    Not read_only because read_only mode doesn't expose tables.
    """
    _validate_file(file_path)
    return openpyxl.load_workbook(str(file_path), read_only=False, data_only=True)


def _validate_file(file_path: str) -> None:
    """Check that file exists and has a supported extension."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise FileValidationError(f"Unsupported file format: {path.suffix}")
